"""The metabolome module: LASSO, split into core and non-core metabolism.

Paper (paper.md:83, 172):

    "we used 33 profiles of 126 metabolites, 53 proteins and 75 genes that
     constitute the core metabolism... We found that using (measured) protein
     expression levels leads to better results (PCC 0.65 +/- 0.21) than by
     using gene expression levels (PCC 0.47 +/- 0.26)...

     For predicting concentrations of metabolites in non-core metabolism, we
     resort to the inference of enzyme concentration from mRNA expression
     levels due to the paucity of profiles with both metabolome (including
     metabolites in non-core metabolism) and proteome information (only 6
     profiles)."

So the split is principled in one direction and a fallback in the other:
  * CORE     from PROTEINS -- established empirically (0.65 vs 0.47). Enzymes,
             not mRNAs, catalyse the reactions that set metabolite pools.
  * NON-CORE from TRANSCRIPTS -- a fallback, because too few profiles pair
             non-core metabolome with proteome. It works acceptably only
             because non-core concentrations have LOW variance
             (0.02 +/- 0.01 vs 0.06 +/- 0.01 for core), which is also why the
             paper's 0.87 must be read against a random baseline of 0.77.

Feature selection follows the same two-track rule as the paper: where
enzyme-substrate relations are known, regress on those enzymes only; otherwise
regress on everything and let the L1 penalty select.
"""

from __future__ import annotations

import re
import warnings
from dataclasses import dataclass, field
from pathlib import Path

import numpy as np

from ecomics.moma._lasso import coefficients, fit_lasso


def load_substrate_enzyme(path=None) -> dict[str, list[str]]:
    """The paper's own enzyme-substrate relations, from Supplementary Data 2.

    The `substrate-enzyme` sheet is two columns -- an enzyme b-number and a
    metabolite KEGG id -- and it is what the Methods mean by "metabolites having
    known enzyme-substrate relations". Returns {metabolite: [enzyme b-numbers]}.

    ⚠ Measure its reach before relying on it: **506 pairs over 317 enzymes but
    only 27 metabolites**, of which **3** appear among the 114 metabolites the
    public release actually ships. So this restricts three regressions and
    leaves 111 to L1, which is a fact about the released table rather than
    about the method.

    ⚠ It is also NOT the core/non-core split. 27 metabolites is not the paper's
    126 core ones, and an earlier note in the diagram atlas claimed this sheet
    would supply that split. It does not; `CORE_KEGG_IDS` stays hand-built.
    """
    import openpyxl

    from ecomics import config as C

    path = Path(path) if path else C.SUPPLEMENTARY["interactions"]
    if not path.exists():
        raise FileNotFoundError(f"{path}\nRun: python scripts/00_acquire.py")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = list(wb["substrate-enzyme"].iter_rows(values_only=True))
    wb.close()

    out: dict[str, list[str]] = {}
    for r in rows:
        if not r or not r[0] or not r[1]:
            continue
        enzyme, metabolite = str(r[0]).strip(), str(r[1]).strip()
        # Skip the header and anything that is not a b-number / C-number pair.
        if not re.fullmatch(r"b\d{4}", enzyme):
            continue
        if not re.fullmatch(r"C\d{5}", metabolite):
            continue
        out.setdefault(metabolite, []).append(enzyme)
    return {m: sorted(set(g)) for m, g in out.items()}

def load_metabolite_enzymes(path=None) -> dict[str, list[str]]:
    """Enzyme-substrate relations from Supplementary Data **1**, not Data 2.

    Two sheets in two files describe the same relation, and they are not
    remotely the same size on the data we actually have:

        Data 2  `substrate-enzyme`  506 pairs, 27 metabolites ->  **3** of our 114
        Data 1  `Metabolite`        451 rows,  212 with enzymes -> **69** of our 114

    `Related enzymes` is a comma-separated b-number list per metabolite, with
    duplicates (ATP names ~100 entries, several twice). Deduplicated and
    intersected with the proteome's 589 measured proteins it leaves a **median
    of 4** enzymes per metabolite -- against 589 columns unrestricted -- and it
    covers **all 12** of the core metabolites the released table contains.

    That is what §3.3.5's "metabolites having known enzyme-substrate relations"
    can actually reach here, so this is the map to pass to
    `MetabolomeModule.fit`. `load_substrate_enzyme` is kept because it is the
    sheet the Methods most directly points at, and because the two disagreeing
    by 23x on reach is itself worth being able to re-measure.

    ⚠ Neither sheet is the core/non-core split. `CORE_KEGG_IDS` stays
    hand-built -- see `load_substrate_enzyme`.
    """
    import openpyxl

    from ecomics import config as C

    path = Path(path) if path else C.SUPPLEMENTARY["metadata"]
    if not path.exists():
        return {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        rows = list(wb["Metabolite"].iter_rows(values_only=True))
    finally:
        wb.close()

    out: dict[str, list[str]] = {}
    for r in rows[1:]:
        # (name, KEGG id, related enzymes) -- the id column carries an
        # abbreviation where no KEGG id exists, which simply will not match a
        # released column and so drops out on its own.
        if not r or len(r) < 3 or not r[1] or not r[2]:
            continue
        met = str(r[1]).strip()
        genes = sorted({g for g in re.split(r"[;,\s]+", str(r[2]))
                        if re.fullmatch(r"b\d{4}", g.strip())})
        if genes:
            out[met] = genes
    return out


__all__ = ["MetaboliteLasso", "MetabolomeModule", "load_substrate_enzyme",
           "load_metabolite_enzymes"]

# Which "N feature columns are empty" messages have already been reported.
#
# `warnings`' own once-per-location dedup does NOT survive this codebase.
# `_lasso.fit_lasso` enters a `warnings.catch_warnings()` block on every fit to
# promote ConvergenceWarning, and entering or leaving that block mutates the
# global filter state -- which invalidates every module's `__warningregistry__`.
# The stdlib's "print this once" is therefore defeated a few thousand times per
# LOCO sweep, and a per-fit warning prints once per fold (50 lines, observed).
#
# Keyed on the whole message rather than a bare flag, so a DIFFERENT count still
# gets through: core trains on 25 conditions and non-core on 6, and "1 of 589
# empty" meaning something different in those two layers is the case worth
# hearing about.
_WARNED_EMPTY_COLUMNS: set[str] = set()


@dataclass
class MetaboliteLasso:
    """Per-metabolite LASSO from upstream molecular levels."""

    alpha: float | None = 1e-3   # fixed: see ProteomeEnsemble's note
    max_features: int = 500
    models: dict[str, object] = field(default_factory=dict, repr=False)
    used: dict[str, np.ndarray] = field(default_factory=dict, repr=False)
    # Failed fits, kept apart from absent ones -- see `_lasso.fit_lasso`.
    errors: dict[str, str] = field(default_factory=dict, repr=False)

    def fit(self, X: np.ndarray, Y: np.ndarray, y_cols: list[str],
            feature_sets: dict[str, np.ndarray] | None = None
            ) -> "MetaboliteLasso":
        """feature_sets: optional metabolite -> column indices of related enzymes.

        Where present, the feature set restricts the regression to known
        enzyme-substrate partners -- prior knowledge instead of selection. Where
        absent, all features are offered and L1 does the selecting.
        """
        n = X.shape[0]
        # Feature columns with no finite value at all. `np.nanvar` reduces those
        # over an empty slice and emits "Degrees of freedom <= 0" -- once per
        # metabolite that reaches the variance ranking, which is where the three
        # unattributed RuntimeWarnings in scripts/04's output came from.
        #
        # The BEHAVIOUR was never wrong: the variance is NaN, `np.argsort` sorts
        # NaN last, so an all-NaN enzyme was already ranked out of the top
        # `max_features`. Only the reporting was. Silencing it with a filter is
        # the one thing not to do here -- these layers run on 6-33 conditions
        # (see the module docstring), so "how much of X is empty" is a fact about
        # whether the fit means anything, not incidental noise.
        #
        # So: detect it ONCE per fit, say how many, and rank the survivors with
        # the all-NaN columns excluded from the reduction rather than suppressed
        # after it. -inf reproduces NaN's sort position exactly, so the selected
        # features are identical to before.
        all_nan = ~np.isfinite(X).any(axis=0)
        if all_nan.any():
            msg = (f"metabolome: {int(all_nan.sum())} of {X.shape[1]} feature "
                   f"columns are entirely non-finite over these {n} samples "
                   f"and carry no variance to rank on; they cannot be "
                   f"selected.")
            if msg not in _WARNED_EMPTY_COLUMNS:   # see the note on the register
                _WARNED_EMPTY_COLUMNS.add(msg)
                warnings.warn(msg, RuntimeWarning)

        for j, met in enumerate(y_cols):
            cols = (feature_sets.get(met) if feature_sets else None)
            if cols is None or len(cols) == 0:
                cols = np.arange(X.shape[1])
            cols = np.asarray(cols)
            if len(cols) > self.max_features:
                keep = ~all_nan[cols]
                var = np.full(len(cols), -np.inf)
                if keep.any():
                    var[keep] = np.nanvar(X[:, cols[keep]], axis=0)
                cols = cols[np.argsort(-var)[:self.max_features]]

            y = Y[:, j]
            ok = np.isfinite(y)
            if ok.sum() < 3:
                continue
            Xd = np.nan_to_num(X[np.ix_(np.flatnonzero(ok), np.asarray(cols))])
            m, err = fit_lasso(Xd, y[ok], alpha=self.alpha)
            if m is None:
                self.errors[met] = err
                continue
            self.models[met] = m
            self.used[met] = np.asarray(cols)
        return self

    def predict(self, X: np.ndarray, y_cols: list[str]) -> np.ndarray:
        out = np.full((X.shape[0], len(y_cols)), np.nan)
        for j, met in enumerate(y_cols):
            m = self.models.get(met)
            if m is None:
                continue
            out[:, j] = m.predict(np.nan_to_num(X[:, self.used[met]]))
        return out

    @property
    def coverage(self) -> int:
        return len(self.models)

    def selected_features(self, met: str, names: list[str]) -> list[tuple[str, float]]:
        """Non-zero coefficients for one metabolite, most important first."""
        m = self.models.get(met)
        if m is None:
            return []
        # `fit_lasso` returns a Pipeline, which has no `coef_` -- so the old
        # `getattr(m, "coef_", np.array([]))` silently returned [] for EVERY
        # metabolite, making "the LASSO selected nothing" indistinguishable
        # from "this method is broken". `coefficients()` reaches the final step.
        coef = coefficients(m)
        idx = np.flatnonzero(np.abs(coef) > 1e-10)
        order = idx[np.argsort(-np.abs(coef[idx]))]
        return [(names[self.used[met][k]], float(coef[k])) for k in order]


@dataclass
class MetabolomeModule:
    """Core (from proteins) and non-core (from transcripts) predictors."""

    core: MetaboliteLasso = field(default_factory=MetaboliteLasso)
    noncore: MetaboliteLasso = field(default_factory=MetaboliteLasso)
    core_metabolites: list[str] = field(default_factory=list)
    noncore_metabolites: list[str] = field(default_factory=list)

    @staticmethod
    def split_core(y_cols: list[str], core_ids: set[str] | None = None
                   ) -> tuple[list[str], list[str]]:
        """Partition metabolites into core and non-core.

        Ecomics identifies metabolites by KEGG C-numbers, so core membership is
        taken from a KEGG-id list of central carbon metabolism (glycolysis,
        pentose phosphate, TCA and immediate branches).
        """
        core_ids = core_ids or CORE_KEGG_IDS
        core = [c for c in y_cols if c in core_ids]
        return core, [c for c in y_cols if c not in core_ids]

    @staticmethod
    def _feature_sets(metabolites: list[str], feature_cols: list[str] | None,
                      enzyme_map: dict[str, list[str]] | None
                      ) -> dict[str, np.ndarray] | None:
        """metabolite -> column indices of its known enzymes, for the ones we have.

        Returns None when nothing can be restricted, which is the signal
        `MetaboliteLasso.fit` reads as "offer all features and let L1 select".
        A metabolite absent from the map, or whose enzymes are all absent from
        `feature_cols`, is simply left out of the dict and gets the full matrix.
        """
        if not enzyme_map or feature_cols is None:
            return None
        pos = {c: i for i, c in enumerate(feature_cols)}
        out: dict[str, np.ndarray] = {}
        for met in metabolites:
            cols = [pos[g] for g in enzyme_map.get(met, ()) if g in pos]
            if cols:
                out[met] = np.asarray(sorted(cols))
        return out or None

    def fit(self, proteins: np.ndarray | None, transcripts: np.ndarray,
            Y: np.ndarray, y_cols: list[str], verbose: bool = False,
            protein_cols: list[str] | None = None,
            transcript_cols: list[str] | None = None,
            enzyme_map: dict[str, list[str]] | None = None
            ) -> "MetabolomeModule":
        """Fit both tracks.

        `enzyme_map` activates the paper's prior-knowledge branch: "For
        metabolites having known enzyme-substrate relations, we predict its
        concentrations from the mRNA expression levels of the related enzymes.
        For those with no such information, we fit from all the genes by using
        LASSO, which allows variable selection." Supply it together with the
        matching `protein_cols` / `transcript_cols` so metabolite names can be
        turned into column indices; `load_substrate_enzyme()` builds it from
        Supplementary Data 2.

        ⚠ It buys less than it sounds on the public release. The sheet names
        **27** metabolites, of which only **3** appear among the 114 released
        columns (2 core, 1 non-core). Everything else falls through to the
        all-features branch, exactly as before. Left optional and defaulting to
        off so the reproduction's headline numbers are unaffected by a change
        that touches three metabolites.
        """
        self.core_metabolites, self.noncore_metabolites = self.split_core(y_cols)
        idx = {c: i for i, c in enumerate(y_cols)}

        if self.core_metabolites and proteins is not None:
            cj = [idx[c] for c in self.core_metabolites]
            fs = self._feature_sets(self.core_metabolites, protein_cols, enzyme_map)
            self.core.fit(proteins, Y[:, cj], self.core_metabolites,
                          feature_sets=fs)
            if verbose:
                print(f"    core    {self.core.coverage:>3d}/"
                      f"{len(self.core_metabolites)} metabolites (from proteins"
                      f"{f', {len(fs)} enzyme-restricted' if fs else ''})")

        if self.noncore_metabolites:
            nj = [idx[c] for c in self.noncore_metabolites]
            fs = self._feature_sets(self.noncore_metabolites, transcript_cols,
                                    enzyme_map)
            self.noncore.fit(transcripts, Y[:, nj], self.noncore_metabolites,
                             feature_sets=fs)
            if verbose:
                print(f"    noncore {self.noncore.coverage:>3d}/"
                      f"{len(self.noncore_metabolites)} metabolites (from transcripts"
                      f"{f', {len(fs)} enzyme-restricted' if fs else ''})")
        return self

    def predict(self, proteins: np.ndarray | None, transcripts: np.ndarray,
                y_cols: list[str]) -> np.ndarray:
        out = np.full((transcripts.shape[0], len(y_cols)), np.nan)
        idx = {c: i for i, c in enumerate(y_cols)}
        if self.core_metabolites and proteins is not None:
            p = self.core.predict(proteins, self.core_metabolites)
            for k, c in enumerate(self.core_metabolites):
                out[:, idx[c]] = p[:, k]
        if self.noncore_metabolites:
            p = self.noncore.predict(transcripts, self.noncore_metabolites)
            for k, c in enumerate(self.noncore_metabolites):
                out[:, idx[c]] = p[:, k]
        return out


# KEGG compound ids for central carbon metabolism: glycolysis, pentose
# phosphate, TCA, and their immediate branch points. This is the "core
# metabolism" the paper distinguishes -- densely measured, high variance across
# conditions, and predicted from proteins rather than transcripts.
#
# ⚠ THIS RECOVERS 12 CORE METABOLITES OF THE PUBLISHED 114; the paper's core set
# is 126 of 356. So "core from protein" against the paper's 0.65 compares a
# 12-metabolite mean against a 126-metabolite one, and `scripts/04` prints the
# two side by side. They are not the same quantity, and the 10x difference in
# set size is a larger caveat than the PCC gap itself.
#
# The list is hand-built from pathway membership because the paper does not
# publish its split, and NO supplementary sheet supplies it -- an earlier version
# of this comment proposed Data 2's "substrate-enzyme" sheet as the way to close
# it, which `load_substrate_enzyme` had already refuted: that sheet names 27
# metabolites, not 126, only 5 of the ids here are in it, and the set is not a
# subset of it or of Data 1 or of their union. 14 ids here appear in neither
# sheet nor the released table, which is what a hand-built list looks like.
#
# ✓ Verified against rest.kegg.jp on 2026-08-17, in both directions: all ids
# resolve to real compounds, each sits in the map it is filed under, and no
# released metabolite that KEGG places in map00010/00020/00030 is missing. The
# reverse direction is the one that found anything -- an omission is invisible
# from the forward check, and a hand-built list fails by omission.
CORE_KEGG_IDS = {
    # glycolysis / gluconeogenesis. Hexose phosphates appear as the generic id
    # AND the beta anomer: KEGG annotates pathways only on the generic entry
    # (C05345 and C05378 carry none), so a table using either id must resolve.
    "C00031", "C00668", "C00085", "C05345", "C00354", "C05378", "C00111",
    "C00118", "C00236", "C00197", "C00631", "C00074", "C00022", "C00186",
    "C00033",
    # pentose phosphate, INCLUDING the Entner-Doudoroff branch: KEGG files
    # C00257 and C04442 under map00030 and E. coli runs edd/eda, so they
    # continue the C01236 -> C00345 steps already here rather than extending
    # scope. Both are in the released 114 -- see the note below on the count.
    "C01236", "C00345", "C00199", "C00117", "C00231", "C00279", "C05382",
    "C00257", "C04442", "C00085",
    # TCA cycle
    "C00024", "C00036", "C00158", "C00417", "C00311", "C00026", "C00091",
    "C00042", "C00122", "C00149",
    # cofactors and energy carriers tightly coupled to the above
    "C00002", "C00008", "C00020", "C00003", "C00004", "C00005", "C00006",
    "C00016", "C00019", "C00035", "C00044",
}
