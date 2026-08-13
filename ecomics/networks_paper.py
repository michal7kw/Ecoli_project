"""The proteome module's networks, read from the paper's OWN Supplementary Data 2.

THE only interaction-graph loader in the repository. It began as a second
source beside `networks.py`, which fetched three of these graphs from live
databases -- RegulonDB via SBRG, STRING v12 at score >= 700, the KEGG REST
API. That comparison ran once, came out +0.113 per-profile PCC in the
supplement's favour, which is why no scraped loader remains.

    networks.py        the Network type, gene_symbol_map, build_cpn
    networks_paper.py  Supplementary Data 2 -- the authors' own edge lists,
                       frozen at publication, already on disk

Why it mattered, since the losing side is gone and cannot demonstrate itself:
both were called "the TRN" and "the PPI" and they were not the same graphs.
The scrape was a 2026 snapshot of databases the paper read in 2016, and the
divergence was large enough to be the dominant term in the proteome layer, whose
entire input is neighbourhood membership.

    network        Data 2      ours     note
    TRN             3,190     7,450     ours is 10 years of RegulonDB later
    PPI           412,236    32,244     12.8x denser
    KEGG           54,788    93,522     ours is a fresh pull, 200-gene cap
    sigma factor    8,180         -     no implementation in networks.py
    small RNA         213         -     no implementation in networks.py

(unique undirected edges.)

COUNT EDGES, NEVER ROWS
-----------------------
The PPI sheet lists every pair in BOTH directions: 823,098 rows collapse to
412,236 undirected edges. Comparing its row count against a deduplicated edge
count on our side inflates the gap to 25x. `Network.n_edges` halves a symmetric
adjacency map, so it is the number to quote on both sides.

WHAT IS NOT LOADED, AND WHY
---------------------------
`substrate-enzyme` (552 rows) maps enzyme b-numbers to KEGG COMPOUND ids
(`b2913 -> C00197`). It is bipartite gene<->metabolite, not a neighbour graph
over genes, so a `NeighbourLassoPredictor` built on it would look up compound
ids in a transcript matrix and find nothing. It belongs to the metabolome
module's substrate mapping, not here. Omitted deliberately, not overlooked.

`CPN` is loaded only on request, under the name `CPN_paper_LEAKY`. See
`load_paper_networks`.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from ecomics import config as C
from ecomics.networks import Network, _add, gene_symbol_map

__all__ = ["PAPER_SHEETS", "LEAKY_CPN", "LEAKY_SUFFIX", "ParseReport",
           "load_paper_networks"]

# Our name -> the sheet it comes from. Names are deliberately plain: the results
# JSON nests them under a `paper` block, so `TRN` there and `TRN` under `ours`
# are unambiguous without a suffix on either.
PAPER_SHEETS = {
    "TRN": "TRN",
    "PPI": "PPI",
    "KEGG": "KEGG pathway",
    "SIGMA": "sigma factor network",
    "SRNA": "small RNA network",
}

# The paper's CPN is a correlation network computed over the FULL compendium, so
# it encodes the very conditions a LOCO fold holds out. The guard is the NAME:
# any arm whose key ends in this suffix is excluded from an ensemble mean by
# `moma.proteome_paper`, and it cannot be reported without the word appearing in
# the output. A flag in a config dict would not survive being copied into a
# table; this does.
LEAKY_SUFFIX = "_LEAKY"
LEAKY_CPN = "CPN_paper" + LEAKY_SUFFIX
_CPN_SHEET = "CPN"

_BNUM = re.compile(r"b\d{4}")
# `ydbA_2`, `yigW_2` -- a symbol carrying a copy index the gene list does not use.
_SUFFIXED = re.compile(r"(.+?)_\d+\Z")

# Rows that are not data. Every sheet has a header; `sigma factor network` also
# carries a TITLE row above it, which is why a naive read of that sheet gives
# 8,382 rather than 8,381: the sheet carries a title row above its header.
_HEADERS = {"tf", "target gene", "gene 1", "gene 2", "srna", "sigma factor",
            "enzyme", "metabolite id (kegg)"}
_TITLE_PREFIX = "supplementary"


@dataclass
class ParseReport:
    """What a sheet contained, including what did not survive parsing.

    `dropped` is reported rather than swallowed because the alternative is a
    loader that quietly returns a smaller graph than the file holds. Fifteen
    distinct tokens across Data 2's data rows are not b-numbers; `_clean`
    recovers 5 and the other 10 are counted and named here.
    """

    sheet: str
    rows: int = 0            # data rows, after title/header
    kept: int = 0            # rows that contributed an edge
    self_loops: int = 0      # `b0020 b0020` -- present in TRN and CPN
    dropped: int = 0         # rows lost to an unmappable token
    dropped_tokens: list[str] = field(default_factory=list)

    def summary(self) -> str:
        return (f"{self.rows} rows -> {self.kept} kept, "
                f"{self.self_loops} self-loops, {self.dropped} dropped")


def _clean(token: object, sym2b: dict[str, str]) -> str | None:
    """One cell -> a b-number, or None if it cannot be one.

    Three repairs, all observed in Data 2 and all cheap:
      `b1474;`  a trailing semicolon
      `citB`    a bare gene symbol
      `ydbA_2`  a symbol with a copy index
    """
    if token is None:
        return None
    t = str(token).strip().rstrip(";").strip()
    if not t:
        return None
    if _BNUM.fullmatch(t):
        return t
    b = sym2b.get(t.lower())
    if b:
        return b
    m = _SUFFIXED.fullmatch(t)
    if m:
        return sym2b.get(m.group(1).lower())
    return None


def _is_data_row(row: tuple) -> bool:
    if not row or row[0] is None:
        return False
    first = str(row[0]).strip().lower()
    return not (first in _HEADERS or first.startswith(_TITLE_PREFIX))


def _network_from_rows(name: str, sheet: str, rows, sym2b: dict[str, str]
                       ) -> tuple[Network, ParseReport]:
    nb: dict[str, set[str]] = {}
    rep = ParseReport(sheet=sheet)
    seen_bad: set[str] = set()
    for row in rows:
        if not _is_data_row(row):
            continue
        rep.rows += 1
        if len(row) < 2:
            rep.dropped += 1
            continue
        a, b = _clean(row[0], sym2b), _clean(row[1], sym2b)
        if a is None or b is None:
            rep.dropped += 1
            for raw, cleaned in ((row[0], a), (row[1], b)):
                if cleaned is None and raw is not None:
                    tok = str(raw).strip()
                    if tok and tok not in seen_bad:
                        seen_bad.add(tok)
                        rep.dropped_tokens.append(tok)
            continue
        if a == b:
            rep.self_loops += 1
            continue
        # `_add` is shared with networks.py on purpose: self-loop policy and the
        # symmetric `set` insert (which de-duplicates the PPI sheet's reciprocal
        # rows) must not diverge between the two sources.
        _add(nb, a, b)
        rep.kept += 1
    return Network(name, nb), rep


def load_paper_networks(path: Path | None = None,
                        include_leaky_cpn: bool = False,
                        strict: bool = True, verbose: bool = True
                        ) -> tuple[dict[str, Network], dict[str, ParseReport]]:
    """The paper's five gene-gene networks, and optionally its leaky CPN.

    include_leaky_cpn  adds the CPN sheet under the key `CPN_paper_LEAKY`.
                       DEFAULT OFF, and off is the only safe default: the paper
                       computed that network across the whole compendium, so
                       every LOCO test fold is inside it. Requesting it is a
                       measurement of what the leak is worth, never an input to
                       a headline number -- `moma.proteome_paper` refuses to
                       average any `*_LEAKY` arm into the ensemble mean.

    strict             raise if any sheet yields ZERO edges. Same reasoning as
                       `networks.load_all_networks`: the failure is silent
                       otherwise, degrading the ensemble to fewer arms with
                       nothing raised and no error to trace.

    Returns (networks, parse reports). The reports are returned rather than
    logged so a caller can put the dropped-token count in `results/*.json`
    beside the numbers it affected.
    """
    path = path or C.SUPPLEMENTARY["interactions"]
    if not path.exists():
        raise FileNotFoundError(
            f"Supplementary Data 2 not found at {path}. "
            f"Run `python scripts/00_acquire.py` to fetch it.")

    wanted = dict(PAPER_SHEETS)
    if include_leaky_cpn:
        wanted[LEAKY_CPN] = _CPN_SHEET

    import openpyxl
    # One handle for every sheet: re-opening per sheet re-parses the shared
    # string table, and the PPI sheet alone is 823k rows.
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sym2b = gene_symbol_map()
    nets: dict[str, Network] = {}
    reports: dict[str, ParseReport] = {}
    try:
        missing = [s for s in wanted.values() if s not in wb.sheetnames]
        if missing:
            raise KeyError(f"{path.name} has no sheet(s) {missing}; "
                           f"it holds {wb.sheetnames}")
        for name, sheet in wanted.items():
            net, rep = _network_from_rows(
                name, sheet, wb[sheet].iter_rows(values_only=True), sym2b)
            nets[name], reports[name] = net, rep
    finally:
        wb.close()

    if verbose:
        for n, net in nets.items():
            print(f"  {n:<16s} {net.n_nodes:>5d} nodes  {net.n_edges:>7d} edges"
                  f"   ({reports[n].summary()})")
    empty = [n for n, net in nets.items() if net.n_edges == 0]
    if strict and empty:
        raise ValueError(
            f"paper network(s) parsed to zero edges: {', '.join(empty)}. "
            f"The sheet layout in {path.name} may have changed. "
            f"Pass strict=False to proceed with a reduced ensemble.")
    return nets, reports
